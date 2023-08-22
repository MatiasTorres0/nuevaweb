from django.shortcuts import render, redirect, get_object_or_404
from .models import Categoria, Producto
from django.core.files.base import ContentFile
from openpyxl import load_workbook
import barcode
from barcode import generate
from barcode.writer import ImageWriter
from io import BytesIO
from .models import Producto, Categoria, Boleta, Conversacion, Mensaje, Tamano, Unidad_medida, VariantePrecio, Ticket, Noticia, Region, Ciudad, Comuna, Promocion
# Create your views here.

def index(request):
    categorias = Categoria.objects.all()
    context = {'categorias': categorias}
    return render(request, 'core/index.html', context)

def base(request):
    categorias = Categoria.objects.all()
    return {'categorias': categorias}


def detalle(request):
    return render(request, 'core/detalle.html')

def contacto(request):
    return render(request, 'core/contacto.html')

def checkout(request):
    return render(request, 'core/checkout.html')

def cart(request):
    return render(request, 'core/cart.html')


def productos_por_categoria(request, categoria_id):
    categoria = Categoria.objects.get(id=categoria_id)
    productos = Producto.objects.filter(categoria=categoria)
    categorias = Categoria.objects.all()
    context = {'categorias': categorias}
    context = {'categoria': categoria, 'productos': productos}
    return render(request, 'core/productos_por_categoria.html', context)



def cargar_categorias_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_categoria = row[0]

            # Crear un nuevo objeto Categoria y guardarlo en la base de datos
            categoria = Categoria(nombre=nombre_categoria)
            categoria.save()

        return redirect('home') 

    return render(request, 'core/cargar_categorias.html')


def cargar_unidades_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_unidad = row[0]

            # Crear un nuevo objeto Unidad_medida y guardarlo en la base de datos
            unidad = Unidad_medida(nombre=nombre_unidad)
            unidad.save()

        return redirect('home') 

    return render(request, 'core/cargar_unidades.html')


def cargar_tamanos_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_tamano = row[0]

            # Crear un nuevo objeto Tamano y guardarlo en la base de datos
            tamano = Tamano(nombre=nombre_tamano)
            tamano.save()

        return redirect('home') 

    return render(request, 'core/cargar_tamanos.html')

def cargar_regiones_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_region = row[0]

            # Crear un nuevo objeto Región y guardarlo en la base de datos
            region = Region(nombre=nombre_region)
            region.save()

        return redirect('home')  # Redirigir a la lista de regiones

    return render(request, 'core/cargar_regiones.html')

def cargar_ciudades_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_ciudad = row[0]
            nombre_region = row[1]
            # Crear un nuevo objeto Ciudad y guardarlo en la base de datos
            ciudad = Ciudad(nombre=nombre_ciudad, region=nombre_region)
            ciudad.save()

        return redirect('home')  # Redirigir a la lista de ciudades

    return render(request, 'core/cargar_ciudades.html')

def cargar_comunas_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_comuna = row[0]
            nombre_ciudad = row[1]
            # Crear un nuevo objeto Comuna y guardarlo en la base de datos
            comuna = Comuna(nombre=nombre_comuna, ciudad=nombre_ciudad)
            comuna.save()

        return redirect('home')  # Redirigir a la lista de comunas

    return render(request, 'core/cargar_comunas.html')

def cargar_productos_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active
        default_categoria = Categoria.objects.get(nombre='Detergente')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_barras = row[0]
            nombre = row[1]
            precio = row[2]
            stock = row[3]
            descuento = row[4] or 0

            # Crear un nuevo objeto Producto y guardarlo en la base de datos
            producto = Producto(
                codigo_barras=codigo_barras,
                nombre=nombre,
                precio=precio,
                stock=stock,
                descuento=descuento,
                categoria=default_categoria
            )
            producto.save()

        return redirect('home') 

    return render(request, 'core/cargar_productos.html')

def crear(request):
    if request.method == 'POST':
        value = request.POST.get('barcode_value', '')

        # Genera el código de barras en formato CODE128
        barcode_class = barcode.get_barcode_class('code128')
        code = barcode_class(value, writer=ImageWriter())
        buffer = BytesIO()
        code.write(buffer)
        code_img = ContentFile(buffer.getvalue())

        context = {'code_img': code_img}
        return render(request, 'core/crear.html', context)

    return render(request, 'core/crear.html')


def creararticulo(request):
    return render(request, "core/creararticulo.html")